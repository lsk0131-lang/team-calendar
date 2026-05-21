import os
import re
from datetime import datetime, timedelta
from collections import defaultdict

from flask import Flask, render_template, jsonify, request, send_file
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build

from calendar_analyzer import fetch_events, parse_event
from media_parser import parse_media, parse_multi_media, is_external_meeting, MEDIA_TIERS, ALL_KEYS as ALL_MEDIA, EXTERNAL_KEYWORDS

app = Flask(__name__)

_INTERNAL_KW = {
    "off", "오프", "휴가", "마이타임", "연차", "반차", "오전반차", "오후반차",
    "병원", "가족돌봄휴가", "돌봄휴가", "가족돌봄", "건강검진", "마타",
    "half", "하프데이",
}

def is_internal_event(mw):
    t = mw.strip().lower()
    return any(t == k or t.startswith(k) for k in _INTERNAL_KW)

CALENDAR_ID = "kakaocorp.com_tg8lp8lctv2v4k7potsjc3fr88@group.calendar.google.com"
TOKEN_FILE   = "token.json"
SCOPES       = ["https://www.googleapis.com/auth/calendar.readonly"]


def get_service():
    creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)
    if creds.expired and creds.refresh_token:
        creds.refresh(Request())
    return build("calendar", "v3", credentials=creds)


# ── 인메모리 캐시 ────────────────────────────────────────────────
_row_cache: dict = {}   # {year: (rows, cached_at)}
CACHE_TTL_CURRENT = timedelta(minutes=10)   # 현재 연도
CACHE_TTL_PAST    = timedelta(hours=1)      # 과거 연도


def load_rows(year=2026, days=None, force=False):
    now = datetime.now()

    # 캐시 확인 (force=False 일 때만)
    if not force and year in _row_cache:
        rows, cached_at = _row_cache[year]
        ttl = CACHE_TTL_CURRENT if year == now.year else CACHE_TTL_PAST
        if now - cached_at < ttl:
            return rows   # 캐시 히트

    service = get_service()
    if days is None:
        start = datetime(year, 1, 1)
        days  = (now - start).days + 1
    events = fetch_events(service, calendar_id=CALENDAR_ID, days_back=days)
    rows = []
    for e in events:
        result = parse_event(e)
        if result:
            rows.extend(result)
    rows = [r for r in rows if r["date"].year == year]

    _row_cache[year] = (rows, now)   # 캐시 저장
    return rows


@app.route("/")
def index():
    return render_template("dashboard.html")

@app.route("/guide")
def guide():
    return render_template("guide.html")


@app.route("/api/sync-media")
def sync_media():
    try:
        from sync_media import sync
        msg = sync()
        # 파서 모듈 재로드 (캐시 초기화)
        import importlib, media_parser
        importlib.reload(media_parser)
        return jsonify({"ok": True, "message": msg})
    except Exception as e:
        return jsonify({"ok": False, "message": str(e)}), 500


@app.route("/api/annual")
def api_annual():
    year  = int(request.args.get("year", 2026))
    month = int(request.args.get("month", 0))   # 0 = 전체
    force = request.args.get("force", "false").lower() == "true"
    quarter = request.args.get("quarter", type=int)
    rows  = load_rows(year=year, force=force)
    if month:
        rows = [r for r in rows if r["date"].month == month]

    # 분기 필터 (특정 월 필터가 없을 때만)
    quarter_months_map = {1:[1,2,3], 2:[4,5,6], 3:[7,8,9], 4:[10,11,12]}
    if quarter and not month:
        q_months = quarter_months_map.get(quarter, [])
        rows = [r for r in rows if r["date"].month in q_months]

    # 기준 기간 계산 (월 단위 threshold 적용용)
    if month:
        months_in_period = 1
    elif quarter and not month:
        now = datetime.now()
        q_months = quarter_months_map.get(quarter, [])
        if year < now.year:
            months_in_period = len(q_months)
        else:
            months_in_period = sum(1 for m in q_months if m <= now.month)
            if months_in_period == 0:
                months_in_period = 1
    else:
        now = datetime.now()
        months_in_period = now.month if year == now.year else 12

    # 티어별 월 기준 임계값: {tier: (good_mo, low_mo)}
    # cnt >= good_mo*months → 충분 / cnt <= low_mo*months → 부족 / 그 사이 → 보통
    # 1티어: 4회↑ 충분, 2~3회 보통, 1회↓ 부족
    # 2티어: 3회↑ 충분, 2회 보통, 1회↓ 부족
    # 3티어: 2회↑ 충분, 1회 보통, 0회 부족
    MONTHLY_THRESHOLD = {1: (4, 1), 2: (3, 1), 3: (2, 0)}

    if not rows:
        return jsonify({"error": f"{year}년 데이터 없음"}), 404

    # ── 월별 집계 ────────────────────────────────────
    if quarter and not month:
        q_months_list = quarter_months_map.get(quarter, [])
        months = [f"{year}-{m:02d}" for m in q_months_list]
    else:
        months = [f"{year}-{m:02d}" for m in range(1, 13)]
    monthly_total   = defaultdict(int)
    monthly_team    = defaultdict(lambda: defaultdict(int))
    monthly_media_t = defaultdict(lambda: defaultdict(int))  # [month][tier]
    cumulative      = []

    for r in rows:
        ym = r["year_month"]
        monthly_total[ym] += 1
        monthly_team[ym][r["team"]] += 1

    cum = 0
    for m in months:
        cum += monthly_total.get(m, 0)
        cumulative.append(cum)

    # ── 매체 분석 ────────────────────────────────────
    # (date, title) 기준으로 중복 제거 → 함께 간 미팅은 1건으로 집계
    media_counts        = defaultdict(int)
    media_counts_lunch  = defaultdict(int)   # 11~14시
    media_counts_dinner = defaultdict(int)   # 17시+
    journalist_list = []
    media_by_month  = defaultdict(lambda: defaultdict(int))
    media_members   = defaultdict(lambda: defaultdict(set))   # [media][member] = set of (date,title)
    media_meetings_raw = {}   # (date, title) -> {date, media, members, journalist}
    seen_media_events      = set()   # (date, title) — 매체 횟수 중복 제거
    seen_journalist_events = set()   # (date, media, journalist) — 기자 횟수 중복 제거

    for r in rows:
        mw = r["meeting_with"]

        # [광고] 등 브라켓 외부 키워드 포함 시 스킵
        if any(tok.strip() in EXTERNAL_KEYWORDS
               for tok in re.findall(r'\[([^\]]+)\]', mw.lower())):
            continue

        # 복수 매체 파싱 — 첫 매체가 미인식이어도 뒤 매체까지 처리
        media_entries = parse_multi_media(mw)
        # 유효 매체만 추출 + 광고·행사 등 외부 키워드 제거
        valid_entries = [
            (m, j, t) for m, j, t in media_entries
            if m and not (j and any(kw in j.lower() for kw in EXTERNAL_KEYWORDS))
        ]
        if not valid_entries:
            continue

        for media, journalist, tier in valid_entries:
            if not media:
                continue

            # 매체 횟수: (날짜 + 원본제목 + 매체) 기준 — 복수 이니셜 중복 제거
            media_event_key = (r["date"], r["title"], media)
            if media_event_key not in seen_media_events:
                seen_media_events.add(media_event_key)
                media_counts[media] += 1
                media_by_month[media][r["year_month"]] += 1
                _h = r.get("start_hour")
                if _h is not None and 11 <= _h < 14:
                    media_counts_lunch[media] += 1
                elif _h is not None and _h >= 17:
                    media_counts_dinner[media] += 1

            # 매체별 미팅 상세 — 툴팁용
            if media_event_key not in media_meetings_raw:
                media_meetings_raw[media_event_key] = {
                    "date": str(r["date"]), "media": media,
                    "members": set(), "journalist": journalist or "",
                    "start_hour": r.get("start_hour"),
                }
            media_meetings_raw[media_event_key]["members"].add(r["member"])

            # 기자 횟수: 콤마 · & · + 구분된 복수 기자를 각각 집계
            journalists = [j.strip() for j in re.split(r'[,&+]', journalist) if j.strip()]
            if not journalists and journalist.strip():
                journalists = [journalist.strip()]
            for j_name in journalists:
                j_key = (r["date"], media, j_name)
                if j_key not in seen_journalist_events:
                    seen_journalist_events.add(j_key)
                    journalist_list.append({
                        "media": media, "journalist": j_name,
                        "date": str(r["date"]), "tier": tier,
                        "members": set(),
                    })
                for entry in reversed(journalist_list):
                    if (entry["media"] == media and entry["journalist"] == j_name
                            and entry["date"] == str(r["date"])):
                        entry["members"].add(r["member"])
                        break

            # 팀원별 매체 미팅 집계
            media_members[media][r["member"]].add((r["date"], r["title"], media))

    # media_members를 건수로 변환
    media_members = {
        media: {member: len(dates) for member, dates in members.items()}
        for media, members in media_members.items()
    }

    # ── 매체별 미팅 상세 (툴팁용) ─────────────────────
    media_meetings_dict = defaultdict(list)
    for info in media_meetings_raw.values():
        members_short = [m.split("(")[0] for m in sorted(info["members"])]
        _h = info.get("start_hour")
        _time = "lunch" if (_h is not None and 11 <= _h < 14) else "dinner" if (_h is not None and _h >= 17) else "other"
        media_meetings_dict[info["media"]].append({
            "date":       info["date"],
            "members":    members_short,
            "journalist": info["journalist"],
            "time":       _time,
        })
    media_meetings_dict = {
        m: sorted(v, key=lambda x: x["date"], reverse=True)
        for m, v in media_meetings_dict.items()
    }

    # ── 기자별 집계 ──────────────────────────────────
    journalist_counts  = defaultdict(int)
    journalist_details = defaultdict(list)
    for j in journalist_list:
        if j["journalist"]:
            key = f"{j['media']} {j['journalist']}"
            journalist_counts[key] += 1
            journalist_details[key].append({
                "members": sorted(j.get("members", set())),
                "date":    j["date"],
            })

    # ── 관리 매체 커버리지 (갭 분석) ────────────────
    from media_parser import load_media_tiers, OTHER_MEDIA
    tiers_map, other_set, aliases_map, all_media = load_media_tiers()

    coverage = {}
    for media in all_media:
        if media not in tiers_map:
            continue
        cnt  = media_counts.get(media, 0)
        tier = tiers_map[media]
        good_mo, low_mo = MONTHLY_THRESHOLD.get(tier, (2, 1))
        good_threshold = good_mo * months_in_period
        low_threshold  = low_mo  * months_in_period
        if cnt >= good_threshold:
            status = "good"
        elif cnt <= low_threshold:
            # low_threshold=0 → 0회 이하이면 부족 (1/2티어)
            # low_threshold=-1 → cnt는 항상 >-1이므로 이 분기 진입 불가 (3티어)
            status = "low"
        elif cnt == 0:
            # 3티어: 0회 → 없음
            status = "none"
        else:
            status = "mid"
        coverage[media] = {
            "count": cnt, "tier": tier,
            "good_threshold": good_threshold,
            "low_threshold":  low_threshold,
            "status": status,
        }

    # 매체별 집계 — 관리 매체(0건 포함) + 기타 매체(실적 있는 것만)
    managed = [
        (media, media_counts.get(media, 0), tiers_map[media])
        for media in all_media if media in tiers_map
    ]
    others = [
        (media, cnt, "기타")
        for media, cnt in media_counts.items()
        if media not in tiers_map and cnt > 0
    ]
    top_media = (
        sorted(managed, key=lambda x: (x[2], -x[1])) +
        sorted(others,  key=lambda x: -x[1])
    )

    # 상위 기자
    top_journalists = sorted(journalist_counts.items(), key=lambda x: -x[1])[:20]

    # 팀원별 매체 커버리지
    member_media = defaultdict(set)
    for r in rows:
        if not is_external_meeting(r["meeting_with"]):
            continue
        media, _, _ = parse_media(r["meeting_with"])
        if media:
            member_media[r["member"]].add(media)

    # ── 분기별 특별 랭킹 ──────────────────────────────
    def month_to_q(m): return (m - 1) // 3 + 1

    lunch_q       = defaultdict(lambda: defaultdict(int))
    dinner_q      = defaultdict(lambda: defaultdict(int))
    other_q       = defaultdict(lambda: defaultdict(int))
    total_q       = defaultdict(lambda: defaultdict(int))
    other_media_q = defaultdict(lambda: defaultdict(list))  # [q][member] = [매체명, ...]
    busy_q        = defaultdict(lambda: defaultdict(int))   # 9-18시 비매체 미팅

    seen_special = set()
    for r in rows:
        key = (r["date"], r["title"], r["member"])
        if key in seen_special:
            continue
        seen_special.add(key)

        mw   = r["meeting_with"].strip()
        q    = month_to_q(r["date"].month)
        hour = r.get("start_hour")   # None이면 종일 이벤트

        # 순수 휴가/연차/off 제외
        if is_internal_event(mw):
            continue

        # [광고] 등 브라켓 외부 키워드 포함 시 스킵
        if any(tok.strip() in EXTERNAL_KEYWORDS
               for tok in re.findall(r'\[([^\]]+)\]', mw.lower())):
            continue

        # 유효 매체 항목
        valid_entries_sr = [
            (m, j, t) for m, j, t in parse_multi_media(mw)
            if m and not (j and any(kw in j.lower() for kw in EXTERNAL_KEYWORDS))
        ]

        if valid_entries_sr:
            # 점심: 11시 이상 14시 미만
            if hour is not None and 11 <= hour < 14:
                lunch_q[q][r["member"]] += 1

            # 저녁: 17시 이상
            if hour is not None and hour >= 17:
                dinner_q[q][r["member"]] += 1

            # 기타 매체 (유효 항목 중 기타 티어만)
            for media, _, tier in valid_entries_sr:
                if tier == "기타":
                    other_q[q][r["member"]] += 1
                    other_media_q[q][r["member"]].append(media)
                    break  # 동일 이벤트 내 기타 매체는 1건만

            # 전체 미팅
            total_q[q][r["member"]] += 1
        else:
            # 비매체 미팅: 9시-18시 사이
            if hour is not None and 9 <= hour < 18:
                busy_q[q][r["member"]] += 1

    def top3(counter):
        return [{"member": m, "count": c}
                for m, c in sorted(counter.items(), key=lambda x: -x[1])[:3]]

    def top3_other(q):
        return [{"member": m, "count": c,
                 "media": sorted(set(other_media_q[q].get(m, [])))}
                for m, c in sorted(other_q[q].items(), key=lambda x: -x[1])[:3]]

    special_rankings = {
        "total":  {str(q): top3(total_q[q])  for q in [1,2,3,4]},
        "lunch":  {str(q): top3(lunch_q[q])  for q in [1,2,3,4]},
        "dinner": {str(q): top3(dinner_q[q]) for q in [1,2,3,4]},
        "other":  {str(q): top3_other(q)     for q in [1,2,3,4]},
        "busy":   {str(q): top3(busy_q[q])   for q in [1,2,3,4]},
    }

    active_members = set(r["member"] for r in rows)
    team_member_counts = defaultdict(set)
    for r in rows:
        team_member_counts[r["team"]].add(r["member"])
    team_member_counts = {k: len(v) for k, v in team_member_counts.items()}

    # media_meetings_raw 직렬화 (set → list, key tuple → string)
    media_meetings_raw_serialized = {
        f"{k[0]}||{k[1]}||{k[2]}": {
            "date": v["date"],
            "media": v["media"],
            "members": [m.split("(")[0] for m in sorted(v["members"])],
            "journalist": v["journalist"],
        }
        for k, v in media_meetings_raw.items()
    }

    _resp = {
        "year": year,
        "quarter": quarter,
        "total": len(rows),
        "member_count": len(active_members),
        "team_member_counts": team_member_counts,
        "months_in_period": months_in_period,
        "months": months,
        "monthly_total": [monthly_total.get(m, 0) for m in months],
        "monthly_teams": {
            team: [monthly_team[m].get(team, 0) for m in months]
            for team in ["기업PR", "서비스PR", "그룹Comm"]
        },
        "cumulative": cumulative,
        "top_media": [{"name": name, "count": cnt, "tier": tier,
                        "lunch": media_counts_lunch.get(name, 0),
                        "dinner": media_counts_dinner.get(name, 0)} for name, cnt, tier in top_media],
        "top_journalists": [
            {"name": k, "count": v, "meetings": sorted([
                {"members": list(e["members"]) if isinstance(e["members"], (set, list)) else e["members"],
                 "date": e["date"]}
                for e in journalist_details[k]
            ], key=lambda x: x["date"], reverse=True)}
            for k, v in top_journalists
        ],
        "coverage": coverage,
        "media_by_month": {
            m: dict(by_month)
            for m, by_month in media_by_month.items()
        },
        "media_meetings": {
            m: [{"date": x["date"], "members": list(x["members"]) if isinstance(x["members"], set) else x["members"], "journalist": x.get("journalist", "")}
                for x in v]
            for m, v in media_meetings_dict.items()
        },
        "media_meetings_raw": media_meetings_raw_serialized,
        "media_members": {
            m: dict(sorted(members.items(), key=lambda x: -x[1]))
            for m, members in media_members.items()
        },
        "tier_summary": {
            str(t): {
                "total": sum(1 for m, info in coverage.items() if info["tier"] == t),
                "good":  sum(1 for m, info in coverage.items() if info["tier"] == t and info["status"] == "good"),
                "mid":   sum(1 for m, info in coverage.items() if info["tier"] == t and info["status"] == "mid"),
                "low":   sum(1 for m, info in coverage.items() if info["tier"] == t and info["status"] == "low"),
                "none":  sum(1 for m, info in coverage.items() if info["tier"] == t and info["status"] == "none"),
            }
            for t in [1, 2, 3]
        },
        "special_rankings": special_rankings,
    }
    # 직렬화 전 set/date 잔류물 제거
    return jsonify(_resp)


@app.route("/api/download")
def download_excel():
    year = int(request.args.get("year", 2026))
    rows = load_rows(year=year)

    thin = Side(style="thin", color="CCCCCC")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    TIER_BG = {"1": "FFE0E0", "2": "FFF8E1", "3": "E8F4FD", None: "FFFFFF"}
    TEAM_BG = {"기업PR": "EDE7F6", "서비스PR": "E8F5E9", "그룹Comm": "FFF3E0", "기타": "F5F5F5"}

    wb = Workbook()

    # ── 시트1: 전체 로데이터 ──────────────────────────────────────
    ws1 = wb.active
    ws1.title = "전체 로데이터"
    headers = ["날짜", "연월", "팀", "이름", "이니셜", "미팅상대", "매체", "기자", "티어", "함께간팀원", "원본제목"]
    widths  = [12,     9,    10,  10,   8,     30,      14,    20,    7,     20,         40]

    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws1.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill = PatternFill("solid", fgColor="2F5496")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = brd
        ws1.column_dimensions[get_column_letter(c)].width = w
    ws1.row_dimensions[1].height = 20
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    seen = set()
    for r in sorted(rows, key=lambda x: x["date"]):
        key = (r["date"], r["title"])
        is_dup = key in seen
        seen.add(key)
        media, journalist, tier = parse_media(r["meeting_with"])
        tier_str = str(tier) if tier else ""
        bg = TEAM_BG.get(r["team"], "FFFFFF")
        vals = [
            str(r["date"]), r["year_month"], r["team"], r["name"], r["initial"],
            r["meeting_with"], media or "", journalist or "", tier_str,
            ", ".join(r.get("together", [])), r["title"],
        ]
        row_idx = ws1.max_row + 1
        for c, val in enumerate(vals, 1):
            cell = ws1.cell(row=row_idx, column=c, value=val)
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="center")
            cell.border = brd
            cell.fill = PatternFill("solid", fgColor=bg)

    # ── 시트2: 매체별 미팅 집계 ──────────────────────────────────
    ws2 = wb.create_sheet("매체별 집계")
    h2 = ["매체", "티어", "총건수", "1월","2월","3월","4월","5월","6월","7월","8월","9월","10월","11월","12월"]
    for c, h in enumerate(h2, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill = PatternFill("solid", fgColor="2F5496")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = brd
    ws2.freeze_panes = "A2"

    from collections import defaultdict
    media_monthly = defaultdict(lambda: defaultdict(int))
    media_tier    = {}
    seen2 = set()
    for r in rows:
        key = (r["date"], r["title"])
        if key in seen2: continue
        seen2.add(key)
        if not is_external_meeting(r["meeting_with"]): continue
        media, _, tier = parse_media(r["meeting_with"])
        if not media: continue
        media_monthly[media][r["date"].month] += 1
        media_tier[media] = tier

    def _tier_num(t): return t if isinstance(t, int) else 9
    for row_idx, (media, monthly) in enumerate(
        sorted(media_monthly.items(), key=lambda x: (_tier_num(media_tier.get(x[0])), -sum(x[1].values()))), 2
    ):
        tier = media_tier.get(media)
        tier_str = str(tier) if tier else "기타"
        total = sum(monthly.values())
        bg = TIER_BG.get(str(tier), "FFFFFF")
        vals = [media, tier_str, total] + [monthly.get(m, 0) for m in range(1, 13)]
        for c, val in enumerate(vals, 1):
            cell = ws2.cell(row=row_idx, column=c, value=val)
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(horizontal="center" if c > 1 else "left", vertical="center")
            cell.border = brd
            cell.fill = PatternFill("solid", fgColor=bg)

    ws2.column_dimensions["A"].width = 18
    for c in range(2, 16):
        ws2.column_dimensions[get_column_letter(c)].width = 8

    # ── 시트3: 팀원별 집계 ───────────────────────────────────────
    ws3 = wb.create_sheet("팀원별 집계")
    h3 = ["팀원", "팀", "총건수", "1월","2월","3월","4월","5월","6월","7월","8월","9월","10월","11월","12월"]
    for c, h in enumerate(h3, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill = PatternFill("solid", fgColor="2F5496")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = brd
    ws3.freeze_panes = "A2"

    member_monthly = defaultdict(lambda: defaultdict(int))
    member_team    = {}
    seen3 = set()
    for r in rows:
        member_team[r["member"]] = r["team"]
        key = (r["date"], r["title"])
        if key in seen3:
            continue
        seen3.add(key)
        mw = r["meeting_with"].strip()
        if is_internal_event(mw):
            continue
        valid = [
            (m, j, t) for m, j, t in parse_multi_media(mw)
            if m and not (j and any(kw in j.lower() for kw in EXTERNAL_KEYWORDS))
        ]
        if not valid:
            continue
        member_monthly[r["member"]][r["date"].month] += 1

    team_order = {"기업PR": 0, "서비스PR": 1, "그룹Comm": 2, "기타": 3}
    for row_idx, (member, monthly) in enumerate(
        sorted(member_monthly.items(), key=lambda x: (team_order.get(member_team.get(x[0]), 9), -sum(x[1].values()))), 2
    ):
        team = member_team.get(member, "")
        total = sum(monthly.values())
        bg = TEAM_BG.get(team, "FFFFFF")
        vals = [member, team, total] + [monthly.get(m, 0) for m in range(1, 13)]
        for c, val in enumerate(vals, 1):
            cell = ws3.cell(row=row_idx, column=c, value=val)
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(horizontal="center" if c > 1 else "left", vertical="center")
            cell.border = brd
            cell.fill = PatternFill("solid", fgColor=bg)

    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 10
    for c in range(3, 16):
        ws3.column_dimensions[get_column_letter(c)].width = 8

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    month = int(request.args.get("month", 0))
    filename = f"PR_미팅데이터_{year}{'_'+str(month)+'월' if month else ''}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    import socket
    ip = socket.gethostbyname(socket.gethostname())
    print(f"🚀 대시보드 실행 중")
    print(f"   내 컴퓨터: http://localhost:5001")
    print(f"   팀원 공유: http://172.20.28.123:5001")
    app.run(host="0.0.0.0", debug=True, port=5001)
